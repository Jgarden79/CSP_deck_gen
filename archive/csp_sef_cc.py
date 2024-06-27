#import standard libraries
import eikon as ek
import math
import datetime
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
pd.options.mode.chained_assignment = None  # default='warn'
import config
import configparser as cp
cfg=cp.ConfigParser()
cfg.read('eikonmjr.cfg.txt')
ek.set_app_key(cfg['eikon']['app_id'])

import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

from pptx import Presentation
from pptx.util import Inches, Pt
import chart_studio
chart_studio.tools.set_credentials_file(username='JGarden79', api_key='eWGoAmjzRp3GIVRTFfSR')
import plotly.io as pio

import plotly.graph_objects as go
from plotly.subplots import make_subplots
from IPython.display import display_html

from PyPDF2 import PdfFileMerger

import options_dat as od

import os

path = 'Images'
isExist = os.path.exists(path) # check if images exists
if not isExist:
     # Create a new directory because it does not exist
    os.makedirs(path)
path_2 = 'html'
isExist = os.path.exists(path_2) # check if images exists
if not isExist:
     # Create a new directory because it does not exist
    os.makedirs(path_2)
#
for f in os.listdir(path):
    os.remove(os.path.join(path,f))

for f in os.listdir(path_2):
    if f != ".git":
        os.remove(os.path.join(path_2,f))

path_3 = 'assets'
for f in os.listdir(path_3):
    os.remove(os.path.join(path_3,f))

def add_disc(file_name_in=list, file_name_out =str ):
    pdfs = []
    for i in file_name_in:
        x = 'images/{}.pdf'.format(i)
        pdfs.append(x)
    pdfs.append('OptionsTemplateSpreadsheetDisclosures.pdf')
    merger = PdfFileMerger(strict=False)
    for pdf in pdfs:
        merger.append(open(pdf, "rb"))
    with open('images/{}.pdf'.format(file_name_out), "wb") as fout:
        merger.write(fout)

def dipsplay_side_by_side(*args):
    html_str = ''
    for df in args:
        html_str += df.to_html()
    display_html(html_str.replace('table', 'table style="display:inline"'), raw=True)


def plot_trade(
    chart_data: pd.DataFrame,
    trade_title: str,
    trade_subtitle: str,
    table_description: str,
    chart_description: str,
    trade_summary_data: list,
    firm: str = 'L'
):
    """
    Creates, formats, and returns a standard plotly figure for a structure option trade given the
    trade data and details.

    Args:
        chart_data (pd.DataFrame): data used to populate the trade return table and chart
        trade_title (str): trade title displayed on the upper-left corner
        trade_subtitle (str): trade subtitle
                Currently used to display the ending date of the trade.
        table_description (str): description placed at the bottom of the trade return data table
                Currently used to describe the trade details and bounds.
        chart_description (str): description placed at the bottom of the trade graph
                Currently used to display the underlying price and timestamp at the time of the
                figure's creation.
        trade_summary_data (list): data displayed in the summary table below the trade title
                Currently displayes the trades structure, term, underlyer, cap, downside before
                protection, and protection level.
        firm (str, optional): firm the trade is being generated for and affects the formatting of
                the resulting figure. Limited to values of 'L' (Lido) or 'O' (Oakhurst). This
                parameter can be updated to an enum in future refactorings.
                Defaults to 'L' (Lido).

    Raises:
        ValueError: if the firm parameter is not one of the handled cases. Allowable values include:
                'L': Lido
                'O': Oakhurst

    Returns:
        plotly.graph_objects._figure.Figure: plotly figure with required formatting that can be
                saved, exported to html, included on a tear-sheet, etc.
    """

    # Set the brand details
    if firm.upper() == 'L':
        # brand settings for lido
        color_scheme = ['#2073C9', '#1A508C', '#041525']
        fonts = ['Noe Display', 'Untitled Sans']
    elif firm.upper() == 'O':
        # brand settings for Oakhurst
        color_scheme = ['#6B8D73', '#4F5151', '#333333']
        fonts = ['Open Sans Light', 'Open Sans Light']
    else:
        _err_str = f"Unhandled firm exception. '{self.firm}' is not a recognized value for firm."
        raise ValueError(_err_str)

    # Create the figure subplots
    fig = make_subplots(cols=3, rows=2, vertical_spacing=0.01, horizontal_spacing=0.05,
                        specs=[
                                [
                                    {'type': 'table', 't': 0.2, 'l': 0.015, },
                                    {'type': 'table', 'colspan': 2, 't': 0.08, 'l': 0, 'r': 0.02},
                                    None,
                                ],
                                [
                                    None,
                                    {'type': 'scatter', 'colspan': 2, 'l': 0.08, 'r': 0.02},
                                    None,
                                ]
                            ])

    # Add the Trade Return
    fig.add_trace(go.Scatter(
        x=chart_data[chart_data.columns[0]],
        y=chart_data[chart_data.columns[1]],
        name='Trade Return',
        mode='lines',
        line=dict(color=color_scheme[0])), row=2, col=2)
    # add the market
    fig.add_trace(go.Scatter(
        x=chart_data[chart_data.columns[0]],
        y=chart_data[chart_data.columns[2]],
        name='Underlying Asset Price Return',
        mode='lines',
        line=dict(color="#1A508C")), row=2, col=2)
    # add the breakeven line
    fig.add_trace(go.Scatter(
        x=chart_data[chart_data.columns[0]],
        y=(chart_data[chart_data.columns[2]] - chart_data[chart_data.columns[2]]),
        name='Break Even',
        mode='lines',
        line=dict(color='red')), row=2, col=2)

    # Format the axes of the graph
    fig.update_xaxes(showgrid=False, gridwidth=1, tickfont={'color': 'black'}, ticklabelposition='inside',
                     title_font={'color': 'black', 'family': fonts[1]})
    fig.update_yaxes(title_text='Return', showgrid=True, gridwidth=1, gridcolor='black', tickformat='.0%', dtick=0.1,
                     range=[chart_data[chart_data.columns[2]].min(), chart_data[chart_data.columns[2]].max()],
                     tickfont={'color': 'black'}, title_font={'color': 'black', 'family': fonts[1]})

    # Set the layout of the figure, titles, descriptions, etc.
    fig.update_layout(
        autosize=False,
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        margin=dict(l=10, r=10, t=10, b=30, autoexpand=False),
        width=816,
        height=1050,
        legend=dict(font={'color': 'black', 'family': fonts[1], 'size': 15, }, x=0.357, y=0.0, orientation='h'),
        title=dict(text=trade_title, font=dict(family=fonts[0], size=31, color='white'), x=0.03, y=0.92),
        annotations=[
            go.layout.Annotation(
                showarrow=False,
                text=trade_subtitle,
                font=dict(family=fonts[1], size=20, color='white'),
                xref='paper',
                yref='paper',
                x=0.04,
                y=0.87),
            go.layout.Annotation(
                showarrow=False,
                text=chart_description,
                font=dict(family=fonts[1], size=12, color='black'),
                xref='paper',
                yref='paper',
                x=0.92,
                y=0.02),
            go.layout.Annotation(
                showarrow=False,
                text=table_description,
                align='left',
                xanchor='left',
                xref='paper',
                x=0.35,
                yanchor='top',
                yref='paper',
                y=0.54,
                font=dict(
                    family=fonts[1],
                    size=13,
                    color="Black")),
        ]
    )

    # Add the data table to the figure
    fig.add_trace(go.Table(
        columnorder=[1, 2, 3],
        columnwidth=[0.333, 0.333, 0.334],
        header=dict(
            values=[f"<b>{i}</b>" for i in list(chart_data.columns)],
            line_color='Black',
            height=20,
            fill_color='white',
            font=dict(size=14, color='black', family=fonts[1])
        ),
        cells=dict(
            values=[
                [f"{i:.2f}" for i in chart_data[chart_data.columns[0]]][0::len(chart_data) // 15],
                chart_data[chart_data.columns[1]][0::len(chart_data) // 15],
                chart_data[chart_data.columns[2]][0::len(chart_data) // 15],
            ],
            fill_color=['white', 'white', 'white'],
            line_color='Black',
            font=dict(size=12, color=['Black', 'Black'], family=fonts[1]),
            height=20,
            align=['center', 'center', 'center'],
            format=[[None], ['.2%'], ['.2%']]
        )
    ), row=1, col=2)

    # Add the trade summary table to the figure
    fig.add_trace(go.Table(
        columnorder=[1, 2],
        columnwidth=[0.33, 0.67],
        header=dict(
            values=['', '<b>Trade</b>'],
            fill_color='rgba(239, 236, 229, 0)',
            line_color='white',
            height=14,
            align=['left', 'center'],
            font=dict(size=11, color='white', family=fonts[1])
        ),
        cells=dict(
            values=[
                [f"<b>{i[0]}:</b>" for i in trade_summary_data],    # catagories
                [f"<b>{i[1]}</b>" for i in trade_summary_data],     # inputs
            ],
            fill_color=['rgba(239, 236, 229, 0)', 'rgba(239, 236, 229, 0)'],
            line_color='white',
            font=dict(size=11, color=['white', 'white'], family=fonts[1]),
            height=14,
            align=['left', 'left']
        )
    ), row=1, col=1)

    fig.update_layout(hovermode="x unified")

    return fig
# define utility functions
def add_disc(file_name_in=list, file_name_out =str ):
    pdfs = []
    for i in file_name_in:
        x = 'images/{}.pdf'.format(i)
        pdfs.append(x)
    pdfs.append('OptionsTemplateSpreadsheetDisclosures.pdf')
    merger = PdfFileMerger(strict=False)
    for pdf in pdfs:
        merger.append(open(pdf, "rb"))
    with open('images/{}.pdf'.format(file_name_out), "wb") as fout:
        merger.write(fout)

def dipsplay_side_by_side(*args):
    html_str = ''
    for df in args:
        html_str += df.to_html()
    display_html(html_str.replace('table', 'table style="display:inline"'), raw=True)

class SEF_and_CovdCalls():

    '''
    Class object to calculate and visualize SEF Strategy (Collar plus synthetic SPY), Covered Calls for Strategic Exit.

    ...

    Parameters
    ..........

    :param RIC: <str> The Reuters identification code for the security in question
    :param get_data: <str> "y" will pull fresh data from Eikon, "n" will use stored data
    :param expiry: <str> Expiration date of the desired options chain
    :param tgt: <int> The target MAX GAIN or MAX LOSS for the collar trade (positive for MAX GAIN, negative for MAX LOSS)
    :param tgt_yld: <flt> The target net cost or net credit of the collar trade (positive for net cost, negative for net credit)
    :param exclude_div: <str> "y" = calculate trade without dividends, "n" = calculate the trade using dividends
    :param custom_div: <str> "y" = use a custom dividend yield (specified when prompted)

    :return: will return output based on the Output param of trade in question


    Attributes
    ..........

        RIC : str
            The Reuters identification code for the security in question

        expiry : str
            Expiration date of the desired options chain

        tgt : int
            IF POSITIVE, gap between current market price and maximum desired gain (non-annualized)
            IF NEGATIVE, gap between current market price and maximum desired gain (non-annualized)

        tgt_yld : int
            Desired net cost of collar trade as a percentage of the underlying security price (non-annualized)
            IF POSITIVE ==> net credit ; IF NEGATIVE ==> net debit

        exclude_div : str
            If 'y', then exlcude expected dividend cash flow as income to be used to pay for net options premium.
            Otherwise, include expected dividend cash flow as income to be used to pay for net options premium.

        custom_div: str
            If "y", then use a custom dividend yield (specified when prompted).

        rng : flt
            Range in percentage of the payoff diagrams

        sym : str
            Ticker symbol of the security in question

        und : df
            DataFrame containing information on the security in question ('Instrument','CF_LAST','YIELD')

        last : flt
            Most recent price of the security in question

        firm : str
            Firm name displayed on PDFs and charts (Lido Advisors or Oakhurst Advisors)

        output : str
            Output style for calc_trade function ('chart', line', or 'rades')

        sorted_trade_chain_calls : df
            Dataframe of imported CALL options data

        sorted_trade_chain_puts : df
            DataFrame of imported PUT options data

        ATM : series
            Series of ATM Call data

        LPUT : series
            Series of Long Put data

        SCALL : series
            Series of Short Call data

        notional : flt
            Notional value of the trade

        inflows : flt
            Income from Short Put premium and expected dividend cash flow (if used)

        tgt_call_val : flt
            Calculated premium required for Short Call

        tgt_put_val : flt
            Calculated premium required for Long Put

        net_opt_cost : flt
            The net premium cost of the selected options (LCALL prem. - SPUT prem.)

        drag : flt
            If dividends are used, he difference between the expected dividend cash flow and the net options premium
            Otherwise, the net options premiumn

        drag_pct : flt
            Drag expressed as a percentage of the underlying stock / ETF price (drag / last)

        cap : flt
            Maximum gain in percent of the strategy

        plug : flt
            Total dividends expected during the duration of the trade. 0 if exclude_div=='y'.

        QTR_div_yld : flt
            Quarterly dividend yield of the underlying security

        months_left : flt
            Time until options expiration in months

        ds_before : flt
            Downside before protection (percent of underlying last)

        upside : flt
            Maximum gain of the trade (percent)

        annual_up : flt
            Annualized maximum gain of the trade (percent)

        downside : flt
            Maximum loss of the trade (percent)

        annual_dn : flt
            Annualized maximum loss of the trade (percent)

        cost : flt
            Total cost of the trade including purchasing the underlying security

    Methods
    ......

    calc_strategy(output,firm)

        Calculates trade parameters for Collar and outputs Pdf, trades or trade line.

        :param output: <str> "line" - will provide terms, "trades" will provide trades, "chart" will provide fact sheet
        :param firm: <str> "L" for Lido, "O" for Oakhurst

    covered_calls()

        Calculates trade parametrs for Covered Calls (strategic exit).

    SEF_payoff_plots()

        Creates SEF strategy payoff diagrams for Collar and SPY Synthetic. Updates Lido SEF Strategy Excel workbok.

    CovdC_all_plots()

        Creates Covered Call and Strategic Exit payoff diagrams. Updates Lido SEF Strategy Excel workbok.


    historical_chart(period='3y')

         Creates historical daily price chart for underlying stock or ETF using non-adjusted close prices from YFinance.
         Intended for use in presentation decks.

         :param period: <str> The lookback period of daily prices ('1y', '3y', '5y', '10y')

    relative_perf_chart(period='3y')

        Credates historical chart of concentrated stock performance vs. SPY performance suing adjusted close prices from YFinance.

    ratio_chart(period='3y')

        Credates ratio chart of concentrated stock performance vs. SPY performance suing adjusted close prices from YFinance.

    '''

    def __init__(self,RIC,get_data,expiry,tgt,tgt_yld,spy_risk,exclude_div,custom_div='n',fee=0.0125,xsp='n'):

        self.RIC=RIC
        self.get_data=get_data
        self.expiry=expiry
        self.tgt=tgt/100
        self.tgt_yld=tgt_yld/100
        self.spy_risk=spy_risk/100
        self.exclude_div=exclude_div
        self.custom_div=custom_div
        self.fee=fee
        self.xsp=xsp

        self.rng=0.4 # range for payoff chart (+/- %)

        if self.get_data == 'y':  # get data if necessary
            od.get_options([self.RIC])
            if self.xsp == 'y':
                od.get_xsp_options()
            else:
                od.get_options(['SPY'])
        else:  # otherwise pass
            pass

        self.sym = pd.read_pickle('assets/{}_sym.pkl'.format(self.RIC)).iloc[0]['ticker']  # import the symbol
        self.und = pd.read_pickle('assets/{}_cached_und.pkl'.format(self.sym))  ##Gets Underlying Price###
        self.last = self.und.iloc[0]['CF_LAST']  # get most recent price        
            
        if self.xsp == 'y':
            self.spy_sym = '.XSP'
            self.spy_dat = pd.read_pickle('assets/{}_cached_und.pkl'.format('XSP'))  ##Gets SPY Price###
            self.spy_last = self.SPY_dat.iloc[0]['CF_LAST']  # get most recent price
        else:        
            self.spy_sym = pd.read_pickle('assets/{}_sym.pkl'.format('SPY')).iloc[0]['ticker']  # import SPY symbol
            self.spy_dat = pd.read_pickle('assets/{}_cached_und.pkl'.format('SPY'))  ##Gets SPY Price###
            self.spy_last = self.spy_dat.iloc[0]['CF_LAST']  # get most recent price

        self.shares = input("Enter number of shares held:")
        self.shares = float(self.shares)
        self.port_val = self.shares * self.last

    def __repr__(self):
        return "SEF_Collar(RIC={},get_data={},expiry={},tgt={},tgt_yld={},spy_risk={},exclude_div={},custom_div={},mgt_fee={},xsp={})".format(
                    self.RIC,
                    self.get_data,
                    self.expiry,
                    self.tgt,
                    self.tgt_yld,
                    self.spy_risk,
                    self.exclude_div,
                    self.custom_div,
                    self.fee,
                    self.xsp)

    def calc_strategy(self,output='line',firm='L'):

        '''Calculates SEF strategy trades'''

        self.output=output
        self.firm=firm

        # get div data
        if self.custom_div == 'n':
            if self.exclude_div == 'n':
                self.div_yld = self.und.iloc[0]['YIELD']
                if type(self.div_yld) == np.float64:
                    self.QTR_div_yld = self.div_yld / 400
                else:
                    self.div_yld = 0
                    self.QTR_div_yld = 0
            else:

                self.div_yld = 0
                self.QTR_div_yld = 0
        else:
            self.QTR_div_yld = input('Enter Annual Dividend Yield:')
            self.QTR_div_yld = float(self.QTR_div_yld) / 4

        self.ann_div_yld = self.div_yld / 100

        # Calculate dividend
        date = datetime.datetime.strptime(self.expiry, '%Y-%m-%d').date()  # convert date to datetime
        today = datetime.datetime.now().date()  # get today's date
        time_left = date - today  # days left
        adj_time_left = time_left / datetime.timedelta(days=1)  # convert to flt
        self.dte = adj_time_left
        adj_time_left_div = np.floor(time_left / datetime.timedelta(days=1) / 365 * 4)
        self.div = np.round(adj_time_left_div * self.QTR_div_yld * self.last, 2)  # total divs expected
        self.div = np.nan_to_num(self.div, 0)  # if no divs replace with 0
        ####################################################################
        # options chain set up for CONCENTRATED STOCK
        chain = pd.read_pickle('assets/{}_cached_chain.pkl'.format(self.sym))  # read in the options chaing to pandas
        chain['MID'] = (chain['CF_BID'] + chain["CF_ASK"]) / 2
        trade_chain = chain[chain['EXPIR_DATE'] == self.expiry]  # isolate the chain we want
        sorted_trade_chain = trade_chain.sort_values(['PUTCALLIND', 'STRIKE_PRC'],
                                                 axis=0, )  # sort the chain by type and strike
        self.sorted_trade_chain_calls = sorted_trade_chain[sorted_trade_chain['PUTCALLIND'] == 'CALL']  # create a call df
        self.sorted_trade_chain_puts = sorted_trade_chain[sorted_trade_chain['PUTCALLIND'] == 'PUT ']  # create a put df

        # ATM CALL
        self.ATM = self.sorted_trade_chain_calls.loc[
                self.sorted_trade_chain_calls['STRIKE_PRC'].sub(self.last).abs().idxmin()]  # get atm call
        self.notional = self.last  # trade's notional is the last traded price of the underlying

        # Solve for individual options legs
        self.tgt_cost = self.last * self.tgt_yld #calcualte target yield (positive = debit, negative = credit)

        if self.tgt < 0: # for target RISK
            self.LPUT = self.sorted_trade_chain_puts.loc[
                            self.sorted_trade_chain_puts['STRIKE_PRC'].sub(self.last * (1 + self.tgt)).abs().idxmin()] # get long put
            self.outflows = self.LPUT['MID'] - self.div # cost of long put minus any dividend income used
            self.SCALL = self.sorted_trade_chain_calls.loc[
                            self.sorted_trade_chain_calls['MID'].sub(self.outflows + self.tgt_cost).abs().idxmin()] # short call
        else: # for target MAX GAIN
            self.SCALL = self.sorted_trade_chain_calls.loc[
                            self.sorted_trade_chain_calls['STRIKE_PRC'].sub(self.last * (1 + self.tgt)).abs().idxmin()] # get short call
            self.inflows = self.SCALL['MID'] + self.div # credit from short call plus any dividend income used
            self.LPUT = self.sorted_trade_chain_puts.loc[
                            self.sorted_trade_chain_puts['MID'].sub(self.inflows - self.tgt_cost).abs().idxmin()] # long put

        self.net_opt_cost = self.LPUT['MID'] - self.SCALL['MID'] # net trade cost (options only)
        self.net_opt_cost_pct = self.net_opt_cost / self.last
        self.cred_deb = np.where(self.net_opt_cost>0,'Debit','Credit')

        self.LPUT['Trans'] = 'BPO'  # add trade type
        self.SCALL['Trans'] = 'SCO'  # add trade type

        Trade = pd.concat([self.SCALL.to_frame().T,
                           self.LPUT.to_frame().T])  # create Trade df
        exp_date = pd.to_datetime(Trade['EXPIR_DATE'].iloc[0]).strftime('%y%m%d')  # get options formated date
        option_type = [s[0] for s in Trade["PUTCALLIND"].to_list()]  # isolate first letter of option type
        strikes = Trade['STRIKE_PRC'].to_list()  # # isolate strikes in list
        option_sym = ['{}{}{}{}'.format(self.sym,
                                        self.expiry,
                                        option_type[i],
                                        int(strikes[i])) for i in range(0, len(Trade))]  # create symbols
        Trade['Symbol'] = option_sym  # add to df
        trade_details = Trade.filter(['Trans', 'Symbol', 'STRIKE_PRC', "MID"])  # create trade_det df
        self.upside = ((self.SCALL['STRIKE_PRC'] - self.net_opt_cost)/ self.last) - 1  # upside calculation
        self.annual_up = self.upside * (365 / adj_time_left)
        self.downside = ((self.LPUT['STRIKE_PRC'] - self.net_opt_cost) / self.last) - 1  # protection calculation
        self.annual_dn = self.downside * (365 / adj_time_left)
        self.months_left = np.round((adj_time_left / 365) * 12, 1)

        self.spread = self.upside - self.downside
        if self.spread > 0.15: #test for constructive sale rule (spread > 15%)
            self.cs = 'PASS'
        else:
            self.cs = 'FAIL'
        trade_line = pd.DataFrame({"Underlying Asset": self.sym,
                                   "Asset Price": self.last,
                                   'Minimum ($)': round(self.notional * 100, -2),
                                   "Expiration Date": Trade['EXPIR_DATE'].iloc[0],
                                   "Months Left": self.months_left,
                                   'Trade Cost' : '{:.2%} {}'.format(abs(self.net_opt_cost_pct),self.cred_deb),
                                   'Potential Upside (%)': '{:.2%}'.format(self.upside),
                                   'Annual Potential Upside (%)': '{:.2%}'.format(self.annual_up),
                                   'Potential Downside (%)': '{:.2%}'.format(self.downside),
                                   'Annual Potential Downside (%)': '{:.2%}'.format(self.annual_dn),
                                   'Spread (%)': '{:.2%}'.format(self.spread),
                                   'Constructive Sale' : self.cs,
                                   'Exclude Div.': '{}'.format(self.exclude_div)},
                                  index=[0])  # creates line

        # begin work on plot data:
        chrt_rng = np.linspace(self.ATM['STRIKE_PRC'] * (1 - self.rng), self.ATM['STRIKE_PRC'] * (1 + self.rng), 50, dtype=float)  # set chart range
        chrt_rng = np.round(chrt_rng, 2)
        scall_ev = [np.maximum(p - self.SCALL['STRIKE_PRC'], 0) * -1 for p in chrt_rng]  # calc scall end val
        lput_ev = [np.maximum(self.LPUT['STRIKE_PRC'] - p, 0) for p in chrt_rng]  # calc lcall end val
        perf_df = pd.DataFrame({"{} Price".format(self.sym): chrt_rng,
                                "SCALL": scall_ev,
                                "LPUT": lput_ev,
                                "UND": chrt_rng,
                                "DIVS": [self.div for i in range(0, len(chrt_rng))]})  # Create the df
        perf_df = perf_df.set_index("{} Price".format(self.sym))  # Set the mkt px as the index
        perf_df['Trade'] = perf_df.sum(axis=1)  # calculate total value
        self.cost = self.LPUT['MID'] - self.SCALL['MID'] + self.last  # total trade cost including underlying
        perf_df['Trade Return'] = ((perf_df['Trade'] - self.net_opt_cost)/ (self.last)) - 1  # trade return
        perf_df = perf_df.sort_index(ascending=False)  # reorder in descending
        perf_df['{} Price Return'.format(self.sym)] = [(p / self.last) - 1 for p in perf_df.index]  # add underlying performance
        rets_tab = perf_df.filter(perf_df.columns[-2:]).reset_index()  # reset index
        fc_date = pd.to_datetime(Trade['EXPIR_DATE'].iloc[0]).strftime("%m-%d-%Y")
        if self.exclude_div == 'n':  # use div data for chart
            self.div_dat = " "
        else:
            self.div_dat = " Excluding Dividends "  # insert if no divs
        self.upside = perf_df['Trade Return'].max()
        self.downside = perf_df['Trade Return'].min()

        ############################################
        #options chain set-up for SPY
        if self.xsp == 'y':
            spy_chain = pd.read_pickle('assets/{}_cached_chain.pkl'.format('XSP'))  # read in the options chaing to pandas
            spy_chain['MID'] = (spy_chain['CF_BID'] + spy_chain["CF_ASK"]) / 2
            spy_trade_chain = spy_chain[spy_chain['EXPIR_DATE'] == self.expiry]  # isolate the chain we want
            spy_sorted_trade_chain = spy_trade_chain.sort_values(['PUTCALLIND', 'STRIKE_PRC'],
                                                     axis=0, )  # sort the chain by type and strike
            self.spy_calls = spy_sorted_trade_chain[spy_sorted_trade_chain['PUTCALLIND'] == 'CALL']  # create a call df
            self.spy_puts = spy_sorted_trade_chain[spy_sorted_trade_chain['PUTCALLIND'] == 'PUT ']  # create a put df            
        else:
            spy_chain = pd.read_pickle('assets/{}_cached_chain.pkl'.format('SPY'))  # read in the options chaing to pandas
            spy_chain['MID'] = (spy_chain['CF_BID'] + spy_chain["CF_ASK"]) / 2
            spy_trade_chain = spy_chain[spy_chain['EXPIR_DATE'] == self.expiry]  # isolate the chain we want
            spy_sorted_trade_chain = spy_trade_chain.sort_values(['PUTCALLIND', 'STRIKE_PRC'],
                                                     axis=0, )  # sort the chain by type and strike
            self.spy_calls = spy_sorted_trade_chain[spy_sorted_trade_chain['PUTCALLIND'] == 'CALL']  # create a call df
            self.spy_puts = spy_sorted_trade_chain[spy_sorted_trade_chain['PUTCALLIND'] == 'PUT ']  # create a put df

        # SPY OTM Put
        self.spy_SPUT = self.spy_puts.loc[
                self.spy_puts['STRIKE_PRC'].sub(self.spy_last * (1 - self.spy_risk)).abs().idxmin()]  # get SPY otm put
        self.spy_LCALL = self.spy_calls.loc[
                self.spy_calls['MID'].sub(self.spy_SPUT['MID']).abs().idxmin()] # SPY long call
        self.spy_net_opt_cost = self.spy_SPUT['MID'] - self.spy_LCALL['MID']

        ###########################################
        # OUTPUT

        if output == 'line':  # line output
            return trade_line
        elif output == 'trades':  # trades output
            self.und['Min_Shares'] = 100
            self.und['Min_Cost'] = self.und['Min_Shares'] * self.last
            print(trade_details)
            print(self.und)
        else:
            print("ERROR")

    def covered_calls(self,cc_output='line',cc_firm='L'):

        '''Calculates covered calls for strategic exit.'''

        self.cc_output=cc_output
        self.cc_firm=cc_firm

        self.cc_expiry = input("Enter Expiration for Covered Calls:")
        self.cc_strike = input("Enter Covered Call Strike Price:")
        self.cc_strike = float(self.cc_strike)
        self.cost_basis = input("Enter Concentrated Stock Cost Basis:")
        self.cost_basis = float(self.cost_basis)

        self.st_rate = 0.37
        self.lt_rate = 0.238

        date = datetime.datetime.strptime(self.cc_expiry, '%Y-%m-%d').date()  # convert date to datetime
        today = datetime.datetime.now().date()  # get today's date
        time_left = date - today  # days left
        adj_time_left = time_left / datetime.timedelta(days=1)  # convert to flt
        self.cc_dte = adj_time_left

        #options chain set-up for Covered Calls
        chain = pd.read_pickle('assets/{}_cached_chain.pkl'.format(self.sym))  # read in the options chaing to pandas
        chain['MID'] = (chain['CF_BID'] + chain["CF_ASK"]) / 2
        trade_chain = chain[chain['EXPIR_DATE'] == self.cc_expiry]  # isolate the chain we want
        sorted_trade_chain = trade_chain.sort_values(['PUTCALLIND', 'STRIKE_PRC'],
                                                 axis=0, )  # sort the chain by type and strike
        self.cc_calls = sorted_trade_chain[sorted_trade_chain['PUTCALLIND'] == 'CALL']  # create a call df

        self.cc_SCALL = self.cc_calls.loc[
                self.cc_calls['STRIKE_PRC'].sub(self.cc_strike).abs().idxmin()] # short call
        self.cc_prem = self.cc_SCALL['MID']
        self.cc_SCALL['Trans'] = 'SCO'  # add trade type
        self.cc_moneyness = self.cc_SCALL['STRIKE_PRC'] / self.last

        Trade = pd.concat([self.cc_SCALL.to_frame().T,
                          ])  # create Trade df
        exp_date = pd.to_datetime(Trade['EXPIR_DATE'].iloc[0]).strftime('%y%m%d')  # get options formated date
        option_type = [s[0] for s in Trade["PUTCALLIND"].to_list()]  # isolate first letter of option type
        strikes = Trade['STRIKE_PRC'].to_list()  # # isolate strikes in list
        option_sym = ['{}{}{}{}'.format(self.sym,
                                        self.cc_expiry,
                                        option_type[i],
                                        int(strikes[i])) for i in range(0, len(Trade))]  # create symbols
        Trade['Symbol'] = option_sym  # add to df
        trade_details = Trade.filter(['Trans', 'Symbol', 'STRIKE_PRC', "MID"])  # create trade_det df
        self.cc_upside = ((self.cc_SCALL['STRIKE_PRC']) / self.last) - 1  # upside calculation
        self.cc_months_left = np.round((self.cc_dte / 365) * 12, 1)

        trade_line = pd.DataFrame({"Underlying Asset": self.sym,
                                   "Asset Price": '${:,.2f}'.format(self.last),
                                   'Minimum ($)': '${:,.0f}'.format(round(self.last * 100, -2)),
                                   "Expiration Date": Trade['EXPIR_DATE'].iloc[0],
                                   "Months Left": self.cc_months_left,
                                   'Premium Received' : '{:.2%} Credit'.format(self.cc_prem / self.last),
                                   'Annualized Premium' : '{:.2%}'.format((self.cc_prem / self.last) * (365/self.cc_dte)),
                                   'Distance to Strike (%)': '{:.2%}'.format(self.cc_upside),
                                   'Max. Gain:' : '{:.2%}'.format(self.cc_upside + (self.cc_prem / self.last))
                                  },
                                  index=[0])  # creates line

        self.und['Min_Shares'] = 100
        self.und['Min_Cost'] = self.und['Min_Shares'] * self.last
        print("**************")
        print("Trade Details:")
        print(trade_details)
        print(self.und)
        print("**************")
        return trade_line

    def SEF_payoff_plots(self):

        '''Generates strategy payoff diagram at expiration for presentation decks'''

        self.synth_pct = input("Enter synthetic replacement percentage:")
        self.synth_pct = float(self.synth_pct) / 100
        self.spy_contracts = math.floor(((self.port_val / self.spy_last) * self.synth_pct) / 100)

        # Functions to calculate options payoffs at expiry
        def call_payoff(stock_range,strike,premium):
            return np.where(stock_range>strike,stock_range-strike,0)-premium
        def put_payoff(stock_range,strike,premium):
            return np.where(stock_range<strike,strike-stock_range,0)-premium

        ########################################
        # CONCENTRATED STOCK #
        # Define stock price range at expiration
        stock_range=np.arange((1-self.rng)*self.last,(1+self.rng)*self.last,1)
        self.mgt_fee = self.fee * (self.dte / 365) * self.port_val

        # Calculate payoffs for individual legs
        long_put=self.LPUT['STRIKE_PRC']
        long_put_prem=self.LPUT['MID']
        short_call=self.SCALL['STRIKE_PRC']
        short_call_prem=self.SCALL['MID']

        payoff_long_put=put_payoff(stock_range,long_put,long_put_prem) * self.shares
        payoff_short_call=call_payoff(stock_range,short_call,short_call_prem) * self.shares * -1
        self.div_cash = self.div * self.shares

        # Calculate Strategy Payoff
        stock_pl=(stock_range-self.last)*self.shares
        strategy_pl=payoff_long_put+payoff_short_call+stock_pl+self.div_cash
        strategy_pl_net=payoff_long_put+payoff_short_call+stock_pl+self.div_cash-self.mgt_fee

        # Create DataFrame of Stock Prices every 5%
        pct_range=1+np.arange(-0.30,0.31,0.05)
        percent_range=pct_range*self.last
        # Caluclate P&L from stock, short call and covered call
        payoff_lp=put_payoff(percent_range,long_put,long_put_prem) * self.shares
        payoff_sc=call_payoff(percent_range,short_call,short_call_prem) * self.shares * -1
        # Calculate Strategy Payoff
        stock_pl_5=(percent_range-self.last)*self.shares
        strat_pl_5=payoff_lp+payoff_sc+stock_pl_5+self.div_cash
        strat_pl_5_net=payoff_lp+payoff_sc+stock_pl_5+self.div_cash-self.mgt_fee
        ret=pct_range-1
        df=pd.DataFrame({'return' : ret,'collar' : strat_pl_5/self.port_val,'collar (net)' : strat_pl_5_net/self.port_val,'stock' : stock_pl_5/self.port_val})
        df.set_index('return',inplace=True)
        df=df.T
        display(df)

        ### Create Visualization
        plt.style.use('fivethirtyeight')
        fig,ax=plt.subplots(figsize=(14,6))
        plt.plot(stock_range,stock_pl,c='grey',lw=3,ls='dashed',label='{} only'.format(self.sym))
        plt.plot(stock_range,strategy_pl,c='green',lw=3,label='{} with Collar (gross)'.format(self.sym))
        plt.plot(stock_range,strategy_pl_net,lw=2,label='{} with Collar (net)'.format(self.sym))
        plt.vlines(x=self.last,ymin=stock_pl.min(),ymax=stock_pl.max(),linestyle='dashed',color='grey',lw=2)
        plt.annotate('Current Stock Price',
                        xy=(self.last,(stock_pl.max()-0)*0.5),
                        fontsize=12,
                        rotation=90,
                        horizontalalignment='right',
                        verticalalignment='center')
        plt.hlines(y=0,xmin=stock_range.min(),xmax=stock_range.max(),color='gray')
        plt.ylabel('Profit / Loss',fontsize=16,fontweight='bold')
        plt.yticks(fontsize=14)
        plt.gca().yaxis.set_major_formatter('${x:,.0f}')
        plt.xlabel('{} Price at Expiration'.format(self.sym),fontsize=16,fontweight='bold')
        plt.xticks(fontsize=14)
        plt.gca().xaxis.set_major_formatter('${x:,.0f}')
        plt.suptitle('Stock: {}     Current Price: ${:.2f}     Option Moneyness: {:.1%} | {:.1%}'.format(self.sym,
                                                                                                         self.last,
                                                                                                         self.LPUT['STRIKE_PRC'] / self.last,
                                                                                                         self.SCALL['STRIKE_PRC']/self.last),
              fontsize=16)
        plt.title('Options expiration: {}'.format(self.expiry),fontsize=14)
        plt.legend(loc='best',fontsize=14)

        # textstr = '\n'.join((
        #     r'Trade Details:',
        #     r'Sell {:.0f} {} {} {} Calls'.format(self.shares/100,self.sym,self.expiry,self.SCALL['STRIKE_PRC']),
        #     r'Buy {:.0f} {} {} {} Puts'.format(self.shares/100,self.sym,self.expiry,self.LPUT['STRIKE_PRC'])))
        # props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
        # ax.text(0.6, 0.25, textstr, transform=ax.transAxes, fontsize=16,
        #         verticalalignment='top', bbox=props)

        plt.tight_layout()
        plt.savefig('images/{}_collar_payoff'.format(self.sym))
        # plt.savefig('images/{}_collar_payoff'.format('BRK-B'))
        plt.show();

        print('Concentrated Stock Value: ${:,.0f}'.format(self.port_val))
        print('{} Last: ${:,.2f}'.format(self.sym,self.last))
        print('SPY Last: ${:,.2f}'.format(self.spy_last))

        # SYNTHETIC EXPOSURE TO SPY #
        # Define price range at expiration
        spy_range=np.arange((1-self.rng)*self.spy_last,(1+self.rng)*self.spy_last,1)
        # Calculate payoffs for individual legs
        short_put=self.spy_SPUT['STRIKE_PRC']
        short_put_prem=self.spy_SPUT['MID']
        long_call=self.spy_LCALL['STRIKE_PRC']
        long_call_prem=self.spy_LCALL['MID']
        payoff_short_put=put_payoff(spy_range,short_put,short_put_prem) * self.spy_contracts * -100
        payoff_long_call=call_payoff(spy_range,long_call,long_call_prem) * self.spy_contracts * 100
        # Calculate Strategy Payoff
        spy_pl = (spy_range - self.spy_last) * self.spy_contracts * 100
        synthetic_pl = payoff_short_put + payoff_long_call
        synthetic_pl_net = synthetic_pl - self.mgt_fee
        ### Create Visualization
        plt.style.use('fivethirtyeight')
        fig,ax=plt.subplots(figsize=(14,6))
        plt.plot(spy_range,spy_pl,c='grey',lw=3,ls='dashed',label='{} only'.format('SPY'))
        plt.plot(spy_range,synthetic_pl,c='green',lw=3,label='{} Synthetic (gross)'.format('SPY'))
        plt.plot(spy_range,synthetic_pl_net,lw=2,label='{} Synthetic (net)'.format('SPY'))
        plt.vlines(x=self.spy_last,ymin=spy_pl.min(),ymax=spy_pl.max(),linestyle='dashed',color='grey',lw=2)
        plt.annotate('Current Stock Price',
                        xy=(self.spy_last,(spy_pl.max()-0)*0.5),
                        fontsize=12,
                        rotation=90,
                        horizontalalignment='right',
                        verticalalignment='center')
        plt.hlines(y=0,xmin=spy_range.min(),xmax=spy_range.max(),color='gray')
        plt.ylabel('Profit / Loss',fontsize=16,fontweight='bold')
        plt.yticks(fontsize=14)
        plt.gca().yaxis.set_major_formatter('${x:,.0f}')
        plt.xlabel('{} Price at Expiration'.format('SPY'),fontsize=16,fontweight='bold')
        plt.xticks(fontsize=14)
        plt.gca().xaxis.set_major_formatter('${x:,.0f}')
        plt.suptitle('Stock: {}     Current Price: ${:.2f}     Option Moneyness: {:.1%} | {:.1%}'.format('SPY',
                                                                                                         self.spy_last,
                                                                                                         self.spy_SPUT['STRIKE_PRC'] / self.spy_last,
                                                                                                         self.spy_LCALL['STRIKE_PRC']/self.spy_last),
              fontsize=16)
        plt.title('Options expiration: {}'.format(self.expiry),fontsize=14)
        plt.legend(loc='best',fontsize=14)
        # textstr = '\n'.join((
        #     r'Trade Details:',
        #     r'Sell {:.0f} {} {} {} Calls'.format(self.SPY_contracts,'SPY',self.expiry,self.SPY_SPUT['STRIKE_PRC']),
        #     r'Buy {:.0f} {} {} {} Puts'.format(self.SPY_contracts,'SPY',self.expiry,self.SPY_LCALL['STRIKE_PRC'])))
        # props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
        # ax.text(0.6, 0.25, textstr, transform=ax.transAxes, fontsize=16,
        #         verticalalignment='top', bbox=props)

        plt.tight_layout()
        plt.savefig('images/SPY_synthetic_payoff_{}'.format(self.sym))
        # plt.savefig('images/SPY_synthetic_payoff_{}'.format('BRK-B'))
        plt.show();

        ## UPDATE SEF STRATEGY ANALYSIS SPREADSHEET ###
        import openpyxl
        xl_file='Lido_CSP_strategy_workbook.xlsx'
        wb=openpyxl.load_workbook(xl_file) #load worksbbok
        ws1=wb.worksheets[0] #define worksheet
        ws1['B4']=self.synth_pct #synthetic replacement percentage
        ws1['B6']=self.sym #concentrated stock symbol
        ws1['B7']=self.last #concentrated stock last
        ws1['B8']=self.ann_div_yld #concentrated stock yield
        ws1['B9']=self.shares #concetnrated stock shares
        if self.exclude_div == 'y':
            ws1['B18']="N"
        elif self.exclude_div == 'n':
            ws1['B18']="Y"
        else:
            ws1['B18']="ERROR"
        ws1['D8']=self.expiry #options expiration date
        ws1['E8']=self.SCALL['STRIKE_PRC'] #short call strike, concentrated stock
        ws1['E9']=self.LPUT['STRIKE_PRC'] #long put strike, concentrated stock
        ws1['I8']=self.SCALL['MID'] #short call premium, concentrated stock
        ws1['I9']=self.LPUT['MID'] #long put premium, concentrated stock

        ws1['B21']=self.spy_last # SPY last
        ws1['B24']=self.spy_contracts # number of SPY contracts
        ws1['E22']=self.spy_LCALL['STRIKE_PRC'] #SPY long call strike
        ws1['E23']=self.spy_SPUT['STRIKE_PRC'] #SPY short put strike
        ws1['I22']=self.spy_LCALL['MID'] #SPY long call premium
        ws1['I23']=self.spy_SPUT['MID'] #SPY short put premium

        wb.save(xl_file)#save workbook
        wb.close()

    def Covd_Call_plots(self):

        '''
        Generates strategy payoff diagram at expiration for presentation decks

        '''
        # Functions to calculate options payoffs at EXPIRY
        def call_payoff(stock_range,strike,premium):
            return np.where(stock_range>strike,stock_range-strike,0)-premium

        # Define stock price range at expiration
        stock_range=np.arange((1 - self.rng) * self.last, (1 + self.rng) * self.last, 1)
        self.cc_mgt_fee = self.fee * (self.cc_dte / 365) * self.port_val

        # Calculate payoffs for individual legs
        short_call=self.cc_SCALL['STRIKE_PRC']
        short_call_prem=self.cc_SCALL['MID']

        payoff_short_call = call_payoff(stock_range,short_call,short_call_prem) * self.shares * -1

        # Calculate Strategy Payoff
        stock_pl = (stock_range - self.last) * self.shares
        strategy_pl = payoff_short_call + stock_pl
        strategy_pl_net = strategy_pl - self.cc_mgt_fee

        # Create DataFrame of Stock Prices every 5%
        pct_range=1+np.arange(-0.30,0.31,0.05)
        percent_range=pct_range*self.last
        # Caluclate P&L from stock, short call and covered call
        payoff_sc=call_payoff(percent_range,short_call,short_call_prem) * self.shares * -1

        # Calculate Strategy Payoff
        stock_pl_5=(percent_range-self.last)*self.shares
        strat_pl_5=payoff_sc+stock_pl_5
        strat_pl_5_net = strat_pl_5 - self.cc_mgt_fee
        ret=pct_range-1

        df=pd.DataFrame({'return':ret,'covered_call':strat_pl_5/self.port_val,'covered_call_net':strat_pl_5_net/self.port_val,'stock':stock_pl_5/self.port_val})
        df.set_index('return',inplace=True)
        df=df.T
        display(df)

        # create visualization
        plt.style.use('fivethirtyeight')
        fig,ax=plt.subplots(figsize=(14,6))
        plt.plot(stock_range,stock_pl,c='grey',lw=3,ls='dashed',label='{} only'.format(self.sym))
        plt.plot(stock_range,strategy_pl,c='green',lw=3,label='{} with Calls (gross)'.format(self.sym))
        plt.plot(stock_range,strategy_pl_net,lw=2,label='{} with Calls (net)'.format(self.sym))
        plt.vlines(x=self.last,ymin=stock_pl.min(),ymax=stock_pl.max(),linestyle='dashed',color='grey',lw=2)
        plt.annotate('Current Stock Price',
                        xy=(self.last,(stock_pl.max()-0)*0.5),
                        fontsize=12,
                        rotation=90,
                        horizontalalignment='right',
                        verticalalignment='center')
        plt.hlines(y=0,xmin=stock_range.min(),xmax=stock_range.max(),color='gray')
        plt.ylabel('Profit / Loss',fontsize=16,fontweight='bold')
        plt.yticks(fontsize=14)
        plt.gca().yaxis.set_major_formatter('${x:,.0f}')
        plt.xlabel('{} Price at Expiration'.format(self.sym),fontsize=16,fontweight='bold')
        plt.xticks(fontsize=14)
        plt.gca().xaxis.set_major_formatter('${x:,.0f}')
        plt.suptitle('Stock: {}     Current Price: ${:.2f}     Option Moneyness: {:.1%}'.format(
                                    self.sym,
                                    self.last,
                                    self.cc_SCALL['STRIKE_PRC']/self.last),
                     fontsize=16)
        plt.title('Options expiration: {}'.format(self.cc_expiry),fontsize=14)
        plt.legend(loc='best',fontsize=14)

        # textstr = '\n'.join((
        #     r'Trade Details:',
        #     r'Sell {:.0f} {} {} {} Calls'.format(self.shares/100,self.sym,self.cc_expiry,self.cc_SCALL['STRIKE_PRC'])))
        # props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
        # ax.text(0.6, 0.25, textstr, transform=ax.transAxes, fontsize=16,
        #         verticalalignment='top', bbox=props)

        plt.tight_layout()
        plt.savefig('images/{}_covd_call_payoff'.format(self.sym))
        # plt.savefig('images/{}_covd_call_payoff'.format('BRK-B'))
        plt.show();
        print('Concentrated Stock Value: ${:,.0f}'.format(self.port_val))

        ## UPDATE PRE-FORMATTED SPREADSHEET ###
        import openpyxl
        xl_file='Lido_CSP_strategy_workbook.xlsx'
        wb=openpyxl.load_workbook(xl_file) #load worksbbok
        ws3=wb.worksheets[2] #define worksheet
        ws3['A4']=self.port_val
        ws3['A6']= self.cc_prem * self.shares #premium collected
        ws3['C6']= (self.cc_prem * self.shares) / self.port_val #pemium collected (%)
        ws3['A8']=(self.cc_SCALL['STRIKE_PRC'] / self.last) - 1 #distance to Call strike
        ws3['A10']='{:.0f} days'.format(self.cc_dte) #days until CC expiration
        ws3['A12']= (self.cc_prem * self.shares) * (365/self.cc_dte)
        ws3['C12']= ((self.cc_prem * self.shares) * (365/self.cc_dte)) / self.port_val
        ws3['A14']= self.cc_SCALL['DELTA']

        ws4=wb.worksheets[3] #define worksheet
        ws4['A3']=f'{self.sym} with Lido CSP (gross)'
        ws4['A4']=f'{self.sym} with Lido CSP (net)'
        ws4['A5']=self.sym
        # #append covered call returns table
        for row in np.arange(0,len(df),1):
            for col in np.arange(0,len(df.columns),1):
                ws4.cell(row=row+3,column=col+2).value=df.iloc[row,col]

        wb.save(xl_file)#save workbook
        wb.close()

        ##############################################################
        ### STRATEGIC EXIT CALCULATION ###
        options_df=pd.DataFrame({'stock':stock_range,'stock_pct':(stock_range/self.last)-1,'options_pl':payoff_short_call})
        options_df.set_index('stock_pct',inplace=True)
        options_df['options_pl_net']=options_df['options_pl'].mul(1-self.st_rate)

        options_df['shrs_to_sell']=np.where(options_df['options_pl']<0,(0-options_df['options_pl']).div((options_df['stock']-self.cost_basis)),
                                            (options_df['options_pl_net']/self.lt_rate)/(options_df['stock']-self.cost_basis))

        options_df['shrs_to_sell'][options_df['shrs_to_sell'] >= self.shares] = self.shares

        options_df['stock_gain']=(options_df['stock'].sub(self.cost_basis)).mul(options_df['shrs_to_sell'])
        options_df['stock_taxes_due']=options_df['stock_gain'].mul(self.lt_rate)
        options_df['pct_shares']=options_df['shrs_to_sell'].div(self.shares)

        ### Strategic Exit Plot ###
        plt.figure(figsize=(12,8))
        plt.style.use('fivethirtyeight')
        plt.title(f'Amount of Liquidation - {self.sym} Tax Neutral Share Sale*',fontsize=18,fontweight='bold')
        # plt.title('Amount of Portfolio Liquidation - Tax Neutral Sale*',fontsize=18,fontweight='bold')

        plt.ylabel('Tax-Neutral Liquidation %',fontsize=16,fontweight='bold')
        plt.ylim(0,options_df['pct_shares'].max()+0.1)
        plt.yticks(fontsize=14)
        plt.gca().yaxis.set_major_formatter('{x:,.0%}')
        plt.xlabel('Stock Price at Expiration',fontsize=16,fontweight='bold')
        plt.xlim(options_df['stock'].min(),options_df['stock'].max())
        plt.xticks(fontsize=14)
        plt.gca().xaxis.set_major_formatter('${x:,.0f}')

        plt.plot(options_df['stock'],options_df['pct_shares'],label='Pct. of Shares to  \nSell "Tax-Neutral"')

        plt.vlines(x=self.last,ymin=0,ymax=options_df['pct_shares'].max()+0.1,linestyle='dashed',color='grey',lw=2)
        # plt.vlines(x=325,ymin=0,ymax=options_df['pct_shares'].max()+0.1,linestyle='dashed',color='red',lw=1)
        # plt.vlines(x=233,ymin=0,ymax=options_df['pct_shares'].max()+0.1,linestyle='dashed',color='red',lw=1)

        plt.annotate('Current Stock Price',
                    xy=(self.last,options_df['pct_shares'].max()*0.7),
                    fontsize=12,
                    rotation=90,
                    horizontalalignment='right',verticalalignment='center')

        x1=np.arange(options_df['stock'].min(),self.cc_SCALL['STRIKE_PRC']+self.cc_SCALL['MID'],0.01) #Below strike
        plt.fill_between(x1,y1=0,y2=options_df['pct_shares'].max()+0.1,color='green',alpha=0.05)

        x2=np.arange(self.cc_SCALL['STRIKE_PRC']+self.cc_SCALL['MID'],options_df['stock'].max(),0.01) #Above Strike
        plt.fill_between(x2,y1=0,y2=options_df['pct_shares'].max()+0.1,color='blue',alpha=0.05)

        plt.legend(loc='upper right',fontsize=16)

        plt.annotate('Stock Flat to Lower',
                    xy=(self.last*0.65,options_df['pct_shares'].max()),
                     fontweight='bold',fontsize=16,
                     horizontalalignment='left',
                     verticalalignment='top')
        plt.annotate('Options make money, \nuse premium to pay tax bill',
                    xy=(self.last*0.65,options_df['pct_shares'].max()*0.925),
                     fontsize=14,
                     horizontalalignment='left',
                     verticalalignment='top')
        plt.annotate('Stock Higher',
                xy=(self.cc_SCALL['STRIKE_PRC']*1.05,options_df['pct_shares'].max()),
                     fontweight='bold',fontsize=16,
                     horizontalalignment='left',
                     verticalalignment='top')
        plt.annotate('Options lose money, \nuse losses to offset gains',
                xy=(self.cc_SCALL['STRIKE_PRC']*1.05,options_df['pct_shares'].max()*0.925),
                     fontsize=14,
                     horizontalalignment='left',
                     verticalalignment='top')
        plt.tight_layout()
        plt.savefig('images/{}_strategic_exit.png'.format(self.sym))
        # plt.savefig('images/{}_strategic_exit.png'.format('BRK-B'))
        plt.show();

    def historical_chart(self,period='3y'):

        '''
        Creates historical daily price chart of underlying stock using non-adjusted close prices from YFinance.
        Intended for use in presentation decks

        :param period: <str> The lookback period of daily prices ('1y', '3y', '5y', '10y')

        '''
        today = datetime.datetime.now().date()  # get today's date
        today_str = today_str=today.strftime('%Y-%m-%d')

        import yfinance as yf
        
        if self.sym == 'BRK.B':
            self.sym = 'BRK-B'
        else:
            pass

        # Underlier Chart
        df=yf.download('{}'.format(self.sym),period=period)['Close'].to_frame()
        plt.style.use('fivethirtyeight')
        plt.figure(figsize=(14,8), linewidth = 5, edgecolor='black')
        plt.plot(df['Close'],label=self.sym)
        plt.tick_params(labeltop=False,labelright=True)
        plt.hlines(y=self.cc_SCALL['STRIKE_PRC'],xmin=df.index[0],xmax=df.index[-1],
                      linestyle='dashed',color='red',alpha=0.5,
                      label='Short Call Strike: ${} (+{:.1%})'.format(self.cc_SCALL['STRIKE_PRC'],(self.cc_SCALL['STRIKE_PRC']/self.last)-1))
        plt.hlines(y=self.last,xmin=df.index[0],xmax=df.index[-1],
                      linestyle='dashed',color='gray',alpha=0.5,
                      label='Current Price: ${}'.format(self.last))
        # plt.hlines(y=self.LPUT['STRIKE_PRC'],xmin=df.index[0],xmax=df.index[-1],
        #               linestyle='dashed',color='green',alpha=0.5,
        #               label='Long Put Strike: ${} ({:.1%})'.format(self.LPUT['STRIKE_PRC'],(self.LPUT['STRIKE_PRC']/self.last)-1))

        plt.title('Stock: {}     |     Last: ${:.2f}     |     Time to First Exp.: {:.0f} days'.format(self.sym,self.last,self.cc_dte),
                    fontsize=16,fontweight='bold')
        plt.legend(loc='upper left',fontsize=14)
        plt.xticks(fontsize=14)
        plt.yticks(fontsize=14)
        plt.gca().yaxis.set_major_formatter('${x:,.0f}')
        plt.xlabel('Date (through {})'.format(today_str),fontsize=14,fontweight='bold')
        plt.ylabel(f'{self.sym} Price',fontsize=16,fontweight='bold')
        plt.tight_layout()
        plt.savefig('images/{}_{}r_price_chart_CovdCall.png'.format(self.sym,period))
        plt.show();

    def relative_perf_chart(self,period='3y'):

        '''
        Creates normalized performance chart of underlying stock versus SPY adjusted close prices from YFinance.
        Intended for use in presentation decks

        :param period: <str> The lookback period of daily prices ('1y', '3y', '5y', '10y')

        '''
        today = datetime.datetime.now().date()  # get today's date
        today_str = today_str=today.strftime('%Y-%m-%d')

        import yfinance as yf

        # Underlier Chart
        df=yf.download([self.sym,'SPY'],period=period)['Adj Close']
        norm=df.div(df.iloc[0]).mul(100).sub(100)

        plt.style.use('fivethirtyeight')
        plt.figure(figsize=(14,8))
        plt.plot(norm[self.sym],label=self.sym)
        plt.plot(norm['SPY'],label='SPY',c='green')

        plt.title('Relative Performance (incl. dividends): {} vs. SPY     |     Period: {}rs'.format(self.sym,period),
                    fontsize=16,fontweight='bold')

        plt.legend(loc='upper left',fontsize=14)
        plt.xticks(fontsize=14)
        plt.yticks(fontsize=14)
        plt.xlabel('Date (through {})'.format(today_str),fontsize=14,fontweight='bold')
        plt.ylabel('Relative Performance (%)',fontsize=16,fontweight='bold')
        plt.tight_layout()
        # plt.savefig('{}_{}r_price_chart.png'.format(self.sym,period))
        plt.show();

    def ratio_chart(self,period='3y'):

        '''
        Creates ratio chart of underlying stock versus SPY adjusted close prices from YFinance.
        Intended for use in presentation decks

        :param period: <str> The lookback period of daily prices ('1y', '3y', '5y', '10y')

        '''
        today = datetime.datetime.now().date()  # get today's date
        today_str = today_str=today.strftime('%Y-%m-%d')

        import yfinance as yf

        # Underlier Chart
        df=yf.download([self.sym,'SPY'],period=period)['Adj Close']

        plt.style.use('fivethirtyeight')
        plt.figure(figsize=(14,8))
        plt.plot(df[self.sym] / df['SPY'],label='Ratio of {} vs. SPY'.format(self.sym))

        plt.tick_params(labeltop=False,labelright=False)

        plt.title('Ratio of {} vs. SPY     |     Period: {}rs'.format(self.sym,period),
                    fontsize=16,fontweight='bold')

        plt.xticks(fontsize=14)
        plt.yticks(fontsize=14)
        plt.xlabel('Date (through {})'.format(today_str),fontsize=14,fontweight='bold')
        plt.ylabel('Ratio',fontsize=16,fontweight='bold')
        plt.tight_layout()
        # plt.savefig('{}_{}r_price_chart.png'.format(self.sym,period))
        plt.show();
